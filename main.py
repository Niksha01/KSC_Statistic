from os import walk
import sys
from openpyxl import load_workbook, Workbook
import os
import psycopg2
import sql
from datetime import date

data = list()
insts = {
    "Детский сад № 1 10.1.3.2": 1,
    "ДС2 (10.4.4.х)": 2,
    "Детский сад № 3 10.1.3.6": 3,
    "Детский сад № 3 10.1.3.6 \\ 10.1.5.66": 3,
    "Детский сад № 4 10.1.3.8": 4,
    "Детский сад № 5 10.1.3.10": 5,
    "ДС6 (10.4.16.х)": 6,
    "Детский сад № 7 10.1.3.14": 7,
    "Детский сад № 8 10.1.3.112": 8,
    "Детский сад № 10 10.1.3.18": 9,
    "Средняя школа № 11 дошколка ДС22 10.1.3.40": 10,
    "Детский сад № 11 10.1.3.20": 10,
    "Детский сад № 12 10.1.3.22": 11,
    "Детский сад №15 10.1.3.28": 12,
    "Детский сад № 17 10.1.3.32": 13,
    "Детский сад № 18 10.1.3.34": 14,
    "Детский сад № 24 10.1.3.42": 15,
    "ДС25 (192.168.х не верная адр)": 16,
    "ДС26 (10.4.64.х)": 17,
    "ДС29 (10.4.70.х)": 18,
    "ДС31 (10.4.75.х)": 19,
    "ДС35 (10.4.81.х)": 20,
    "ДС37 (10.4.87.х)": 21,
    "Детский сад №39 10.1.3.64": 22,
    "Гимназия № 39 дошколка ДС41 10.1.3.68": 66,
    "Детский сад № 40 10.1.3.66": 23,
    "Детский сад № 42 10.1.3.70": 24,
    "Детский сад № 43 10.1.3.72": 25,
    "ДС44 (10.4.108.х)": 26,
    "ДС45 (10.4.111.х)": 27,
    "ДС46 (10.4.114.х)": 28,
    "ДС47 (10.4.117.х)": 29,
    "ДС48 (10.4.120.х)": 30,
    "ДС50 (10.4.123.х)": 31,
    "ДС51 (10.4.126.х)": 32,
    "ДС53 (10.4.129.х)": 33,
    "ДС56 (10.4.132.х)": 34,
    "ДС58 (10.4.138.х)": 35,
    "ДС63 (10.4.141.х)": 36,
    "ДС70 (10.4.168.х)": 37,
    "ДС72 (10.4.174.х)": 38,
    "СШ1 (10.41.32.х)": 39,
    "Средняя школа № 3 10.1.2.6": 40,
    "СШ4 (10.41.63.х)": 41,
    "СШ5 (10.41.51.х)": 42,
    "СШ6 (10.41.52.х)": 43,
    "СШ7 (10.41.66.х)": 44,
    "Средняя школа №8 10.1.2.16": 45,
    "Средняя школа № 9 10.1.2.18": 46,
    "СШ10 (10.41.53.х)": 47,
    "СШ11 (10.41.54.х) есть не верная адр": 48,
    "СШ12 (10.41.55.х)": 49,
    "СШ15 (10.41.56.х)": 50,
    "СШ17 (10.41.57.х)": 51,
    "Средняя школа № 17 дошколка ДС9 10.1.3.16": 51,
    "Средняя школа № 20 10.1.2.38": 52,
    "Лицей № 21 (10.41.47.х)": 53,
    "Средняя школа № 24 дошколка ДС19 10.1.3.36": 54,
    "СШ26 (10.41.60.х)": 55,
    "СШ27 (10.41.34.х)": 56,
    "Средняя школа № 28 10.1.2.50": 57,
    "СШ30 (10.41.37.х)": 58,
    "Средняя школа № 31 10.1.2.54": 59,
    "СШ32 (10.41.49.х)": 60,
    "Средняя школа №33 10.1.2.58": 61,
    "Средняя школа № 34 10.1.2.60": 62,
    "СШ35 (10.41.62.х)": 63,
    "СШ36 (10.41.40.х)": 64,
    "СШ37 (10.41.50.х)": 65,
    "Гимназия № 39 (10.3.102.х)": 66,
    "Средняя школа № 40 10.1.2.72": 67,
    "СШ41 (10.41.65.х)": 68,
    "Средняя школа № 42 10.1.2.76": 69,
    "СШ43 (10.41.42.х)": 70,
    "Средняя школа № 45 10.1.2.80": 71,
    "Лицей № 46 (10.41.48.х)": 72,
    "СШ52-ДС52 (10.41.31.х)": 73,
    "Начальная сад-школа № 52 10.1.2.84": 73,
    "ДДТ Юность (10.4.156.х)": 74,
    "ЦВР (10.4.162.х)": 75,
    "СЮТ (10.4.159.х)": 76,
    "ЦТРиГО (10.3.138.х)": 77,
    "ДЮСШ1 (10.3.165.х)": 78,
    "ДЮСШ2 (10.3.147.х)": 79,
    "ДЮСШ3 (10.3.135.х)": 80,
    "ДЮСШ4 (10.3.153.х)": 81,
    "Сервер KSC-Z": 83,
    "ДС16 (10.4.43.x)": 56,
    "ДС20 (10.4.55.х)": 49,
    "ДС30 (2-е здание дс38) (10.4.72.х)": 39,
    "ДС41": 66
}


# получаем путь файла
def get_filepath():
    current_dir = os.path.dirname(os.path.abspath(__file__))  # Определяем текущую директорию
    # current_dir = os.path.dirname(sys.executable)  # Определяем текущую директорию
    print(current_dir)
    mypath = os.path.join(current_dir, 'sheet')
    filename = next(walk(mypath), (None, None, []))[2]
    if len(filename) > 1:
        raise ValueError("В папке sheet содержится больше одного файла. Пожалуйста, удалите лишние файлы")
    elif len(filename) == 0:
        raise ValueError("Отсутствует файл в папке sheet")
    else:
        filename = filename[0]
    file_path = os.path.join(mypath, filename)

    # Проверяем формат файла
    if filename.endswith('.xls'):
        raise ValueError("Файл имеет неподдерживаемый формат: .xls. Пожалуйста, используйте файл .xlsx.")
    elif filename.endswith('.xlsx'):
        return file_path  # Возвращаем оригинальный файл, если он уже в формате .xlsx
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {filename}")


# открываем таблицу с отчетом
def open_sheet(path):
    return load_workbook(path)


def connect_to_db() -> object:
    connection = psycopg2.connect(
        host="192.168.10.195",
        port="5432",
        database="KSCStat",
        user="postgres",
        password="root"
    )
    connection.autocommit = True
    cursr = connection.cursor()
    return cursr


def record_insts_to_db(workbook, cursor):
    insts = [
        "МАДОУ «Детский сад № 1 комбинированного вида»",
        "МАДОУ «Центр развития ребёнка – детский сад № 2»",
        "МАДОУ «Детский сад № 3 комбинированного вида»",
        "МАДОУ «Детский сад № 4 комбинированного вида»",
        "МБДОУ «Детский сад № 5 комбинированного вида»",
        "МАДОУ «Детский сад № 6 комбинированного вида»",
        "МАДОУ «Детский сад № 7 комбинированного вида»",
        "МАДОУ «Центр развития ребенка - детский сад № 8»",
        "МБДОУ «Детский сад № 10 комбинированного вида»",
        "МАДОУ «Детский сад № 11 комбинированного вида»",
        "МБДОУ «Детский сад № 12 присмотра и оздоровления»",
        "МБДОУ «Детский сад № 15 комбинированного вида»",
        "МАДОУ «Детский сад № 17 общеразвивающего вида»",
        "МБДОУ «Детский сад № 18 общеразвивающего вида»",
        "МБДОУ «Детский сад № 24 общеразвивающего вида»",
        "МАДОУ «Детский сад № 25»",
        "МБДОУ «Детский сад № 26 общеразвивающего вида»",
        "МАДОУ «Детский сад № 29 комбинированного вида»",
        "МБДОУ «Детский сад № 31 комбинированного вида»",
        "МБДОУ «Детский сад № 35»",
        "МБДОУ «Детский сад № 37 комбинированного вида»",
        "МАДОУ «Центр развития ребёнка – детский сад № 39»",
        "МБДОУ «Детский сад № 40 комбинированного вида»",
        "МАДОУ «Детский сад № 42 комбинированного вида»",
        "МАДОУ «Детский сад № 43 - Центр развития ребёнка»",
        "МБДОУ «Детский сад № 44»",
        "МБДОУ «Детский сад № 45 общеразвивающего вида»",
        "МАДОУ «Детский сад № 46 комбинированного вида»",
        "МБДОУ «Детский сад № 47 общеразвивающего вида»",
        "МБДОУ «Детский сад № 48 комбинированного вида»",
        "МАДОУ «Детский сад № 50 комбинированного вида»",
        "МАДОУ «Детский сад № 51 комбинированного вида»",
        "МБДОУ «Детский сад № 53 общеразвивающего вида»",
        "МАДОУ «Детский сад № 56 комбинированного вида»",
        "МАДОУ «Детский сад № 58 комбинированного вида»",
        "МБДОУ «Детский сад № 63 общеразвивающего вида»",
        "МАДОУ «Детский сад № 70»",
        "МАДОУ «Детский сад № 72»",
        "МАОУ «Средняя школа № 1»",
        "МАОУ «Средняя школа № 3 имени А.С. Пушкина»",
        "МБОУ «Средняя школа № 4»",
        "МБОУ «Основная школа № 5»",
        "МБОУ «Основная школа № 6»",
        "МБОУ «Средняя школа № 7»",
        "МАОУ «Средняя школа № 8»",
        "МБОУ «Средняя школа № 9»",
        "МБОУ «Средняя школа № 10»",
        "МБОУ «Средняя школа № 11 им. В. Д. Бубенина»",
        "МБОУ «Средняя школа № 12»",
        "МБОУ «Средняя школа № 15»",
        "МБОУ «Средняя школа № 17 им. В.С. Завойко»",
        "МБОУ «Средняя школа № 20»",
        "МБОУ «Лицей № 21»",
        "МАОУ «Средняя школа № 24»",
        "МБОУ «Средняя школа № 26»",
        "МАОУ «Средняя школа № 27»",
        "МАОУ «Средняя школа № 28 им. Г.Ф. Кирдищева»",
        "МАОУ «Средняя школа № 30»",
        "МАОУ «Средняя школа № 31»",
        "МБОУ «Основная школа № 32»",
        "МАОУ «Средняя школа № 33»",
        "МБОУ «Средняя школа № 34»",
        "МБОУ «Средняя школа № 35»",
        "МАОУ «Средняя школа № 36 имени П.Т. Новограбленова»",
        "МБОУ «Основная школа № 37»",
        "МАОУ «Гимназия № 39»",
        "МБОУ «Средняя школа № 40»",
        "МБОУ «Средняя школа № 41»",
        "МАОУ «Средняя школа № 42»",
        "МАОУ «Средняя Школа № 43»",
        "МАОУ «Средняя школа № 45»",
        "МБОУ «Лицей № 46»",
        "МАОУ «Начальная школа - детский сад № 52»",
        "МБОУ ДО «ДДТ «Юность»",
        "МБУДО «Центр внешкольной работы»",
        "МБУ ДО «Станция детского и юношеского технического творчества»",
        "МБУ ДО «ЦТРиГО»",
        "МБУДО «Спортивная школа № 1»",
        "МАУДО «Спортивная школа № 2»",
        "МБУДО «Спортивная школа № 3»",
        "МБУДО «Спортивная школа № 4»",
        "МБУДО «Спортивная школа № 5»",
        "МАУ «ИМЦ»"
    ]

    i = 0

    for inst in insts:
        cursor.execute(f"insert into institutions( institution_name) values( '{inst}')")
        i += 1


def record_keys_to_bd(workbook, cursor):
    print("Загрузка ключей")
    data = list()
    sheet = workbook['Summary']
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, values_only=True):
        data.append([row[0], row[3], row[4]])
    """   
    for i in range(0, len(data)):
        cursor.execute(f"insert into keys_ksc(key, resource, validity_period)" +
                      f"'{data[i][0]}', {int(data[i][1])}, CAST('{data[i][2].date()}' AS DATE))")
    """

    # Проходим по каждому элементу в data
    cnt = 0
    for el in data:
        # Проверяем, есть ли запись в базе данных
        cursor.execute(
            f"SELECT COUNT(*) FROM keys_ksc WHERE key = '{el[0]}'"
        )
        is_in_db = cursor.fetchone()

        if is_in_db[0] > 0 or None in el:
            # print(f"Запись уже существует в базе данных: {el}")
            continue
        # Добавляем новую запись в таблицу
        cursor.execute(f"insert into keys_ksc(key, resource, validity_period)" +
                       f"values('{el[0]}', {int(el[1])}, "
                       f"CAST('{el[2].date()}' AS DATE))")
        cnt += 1
        print(f"Добавлена новая запись: {el}")
    print(f'Добавлено {cnt} записей')
    # Получаем все записи из базы данных
    cursor.execute("SELECT key FROM keys_ksc")
    db_records = {row[0] for row in cursor.fetchall()}

    # Выводим записи, которые есть в базе, но отсутствуют в data
    data_key = set()
    for d in data:
        data_key.add(d[0])
    missing_in_data = db_records - data_key
    if missing_in_data:
        print("Следующие записи есть в базе данных, но отсутствуют в новых данных:")
        for record in missing_in_data:
            print(f"- {record}")
    else:
        print("Все записи из базы данных были учтены.")


def record_insts_ksc_to_db(workbook, cursor):
    print("Загрузка учреждений")
    # Собираем уникальные данные из двух листов Excel
    data = set()
    sheet = workbook['Details']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[1] != None:
            data.add(row[1])

    sheet = workbook['Slave servers summary']
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        if row[0] is not None and "Подчиненный Сервер" in row[0]:
            data.add(row[0].replace('Подчиненный Сервер администрирования "Управляемые устройства : ', '')
                     .replace("\"", "")
                     .replace(" недоступно", ""))

    # Получаем текущий максимальный идентификатор в таблице
    cursor.execute("SELECT COALESCE(MAX(institution_ksc_id), 0) FROM institutions_ksc")
    current_max_id = cursor.fetchone()[0]
    next_id = current_max_id + 1  # Начинаем с первого доступного ID
    cnt = 0
    # Проходим по каждому элементу в data
    for el in data:
        # Проверяем, есть ли запись в базе данных
        cursor.execute(
            f"SELECT COUNT(*) FROM institutions_ksc WHERE institution_ksc_name = '{el}'"
        )
        is_in_db = cursor.fetchone()[0] > 0

        if is_in_db:
            #print(f"Запись уже существует в базе данных: {el}")
            continue

        # Проверяем, есть ли элемент в словаре insts
        if el in insts:
            institution_id = insts[el]

            # Добавляем новую запись в таблицу
            cursor.execute(
                "INSERT INTO institutions_ksc (institution_ksc_name, institution_id) VALUES (%s, %s)",
                (el, institution_id)
            )
            cnt += 1
            # Получаем имя учреждения
            cursor.execute("SELECT institution_name FROM institutions WHERE institution_id = %s", (institution_id,))
            added_institution = cursor.fetchone()[0]

            print(f"Добавлена новая запись: {el}, учреждение: {added_institution}")
            next_id += 1
        else:
            print(f"{el} добавленна в БД, введите id вручную")
            # Добавляем новую запись в таблицу
            cursor.execute(
                "INSERT INTO institutions_ksc (institution_ksc_name, institution_id) VALUES (%s, %s)",
                (el, None)
            )
    print(f'Добавлено {cnt} записей')
    # Получаем все записи из базы данных
    cursor.execute("SELECT institution_ksc_name FROM institutions_ksc")
    db_records = {row[0] for row in cursor.fetchall()}

    # Выводим записи, которые есть в базе, но отсутствуют в data
    missing_in_data = db_records - data
    if missing_in_data:
        print("Следующие записи есть в базе данных, но отсутствуют в новых данных:")
        for record in missing_in_data:
            print(f"- {record}")
    else:
        print("Все записи из базы данных были учтены.")


def record_key_inst_relation(workbook, cursor):
    print("Загрузка связей ключ-учреждение")
    data = set()
    sheet = workbook['Details']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[1] is not None and row[5] is not None:
            data.add(row[1] + "#" + row[5])
    current_institution = ""
    sheet = workbook['Slave servers summary']
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, values_only=True):
        if row[0] is None or row[0] == "Лицензионный ключ":
            continue
        if "Подчиненный Сервер администрирования" in row[0]:
            current_institution = (row[0].replace('Подчиненный Сервер администрирования "Управляемые устройства : ', '')
                                   .replace("\"", "")
                                   .replace(" недоступно", ""))
            continue
        license_key = row[0]
        data.add(current_institution + "#" + license_key)

    cursor.execute("""
            SELECT CONCAT(i.institution_ksc_name, '#', k.key) AS record
            FROM institutions_ksc_keys_ksc ik
            JOIN institutions_ksc i ON ik.institution_ksc_id = i.institution_ksc_id
            JOIN keys_ksc k ON ik.key_ksc_id = k.key_ksc_id
        """)

    db_data = set(row[0] for row in cursor.fetchall())
    missing_in_data = db_data - data
    data = data - db_data
    cnt = 0
    for record in data:
        values = record.split('#')
        # Проверка institution_ksc_id
        cursor.execute("SELECT institution_ksc_id FROM institutions_ksc WHERE institution_ksc_name = %s", (values[0],))
        institution_result = cursor.fetchone()
        if institution_result is None:
            continue
        institution_ksc_id = institution_result[0]

        # Проверка key_ksc_id
        cursor.execute("SELECT key_ksc_id FROM keys_ksc WHERE key = %s", (values[1],))
        key_result = cursor.fetchone()
        if key_result is None:
            continue
        key_ksc_id = key_result[0]

        # Вставка данных
        cursor.execute(
            "INSERT INTO institutions_ksc_keys_ksc (institution_ksc_id, key_ksc_id) "
            "VALUES (%s, %s)", (institution_ksc_id, key_ksc_id)
        )
        cnt += 1
    print(f'Добавлено {cnt} записей')
    if len(missing_in_data) != 0:
        print("Следующие записи есть в базе данных, но отсутствуют в новых данных:")
        for record in missing_in_data:
            print("- " + record)
    else:
        print("Все записи из базы данных были учтены.")


def record_keys_ksc_usage(workbook, cursor):
    sheet = workbook['Summary']
    condition_check = True
    while condition_check:
        report_date = input("Введите дату отчета в формате dd.mm.yyyy")
        try:
            cursor.execute("SELECT file_date FROM keys_ksc_usage WHERE keys_ksc_usage.file_date = CAST(%s AS DATE)",
                           (report_date,))
            if len(report_date) == 10 and (cursor.fetchone() == [] or cursor.fetchone() == None):
                condition_check = False
            else:
                print("Отчет на заданную дату уже существует")
        except Exception:
            print("Неверный формат даты")

    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, values_only=True):

        if row[0] is None:
            continue
        cursor.execute(f"SELECT key_ksc_id FROM keys_ksc WHERE keys_ksc.key = %s",
                       (row[0],))
        value = cursor.fetchone()
        if not value is None or value != []:
            cursor.execute(f"INSERT INTO keys_ksc_usage(key_ksc_id, key_usage, file_date)"
                           f"VALUES(%s, %s, CAST(%s AS DATE))", (value, row[1], report_date))
        else:
            print(value + "not found!")


if __name__ == '__main__':
    # Получаем путь к файлу
    file_path = get_filepath()
    print(f"Имя файла: {file_path}")

    # Открываем и читаем файл .xlsx
    workbook = open_sheet(file_path)

    cursor = connect_to_db()
    cursor.execute(sql.sql)
    record_insts_to_db(workbook, cursor)
    record_keys_to_bd(workbook, cursor)
    record_insts_ksc_to_db(workbook, cursor)
    record_key_inst_relation(workbook, cursor)
    record_keys_ksc_usage(workbook, cursor)
